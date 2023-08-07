from openpyxl import load_workbook

Op = []
On =[]
Ap = []
An = []
Bp =[]
Bn = []
ABp = []
ABn = []

filename = "E:\\ADS PROJECT PYTHON\\Program files\\Donor_details.xlsx"
wb = load_workbook(filename)
sh1 = wb['Donor Details']
a = sh1.max_row
b = sh1.max_column

def sync():
    for i in range(2, a+1):
        ag = sh1.cell(i, 3).value
        bg = globals()[ag]
        bg.append(sh1.cell(i, 1).value)



def disp(var):
    return var[0]

sync()
print(ABp)



