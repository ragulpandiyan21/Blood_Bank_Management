import tkinter as Tk
from cProfile import label
from tkinter import *
from PIL import Image,ImageTk
from tkinter import messagebox
from openpyxl import load_workbook

root = Tk()
root.title("Blood Bank Mgmt")
w=1200
h=750
sw = root.winfo_screenwidth()
sh = root.winfo_screenheight()
x = (sw/2) - (w/2)
y = (sh/2) - (h/2)
root.geometry("%dx%d+%d+%d" % (w,h,x,y))

#font style
look = "Adobe Caslon Pro"

######booking details
booking_frame = Frame(root, bg="white", width=900, height=600).place(x=150, y=50)
title = Label(booking_frame, text= "DONOR DETAILS SEARCHING",font= ("look", 20),fg= "red", background = "white")
title.place(x=150, y=90)

####Init search
id_lbl=Label(booking_frame, text="Donor Id",font= ("look", 12), bg = "white")
id_lbl.place(x=170,y=150)
id_lnlE=Entry(booking_frame,font=("look","12"), bg="light blue", bd=1)
id_lnlE.place(x=350, y=150, width=200, height=40)


def search_DD():
    id_da = id_lnlE.get()
    filename = "E:\\ADS PROJECT PYTHON\\Program files\\Donor_details.xlsx"
    wb = load_workbook(filename)
    sh1 = wb['Donor Details']
    sh2 = wb['START DID']
    max = sh1.max_row
    try:
        int(id_lnlE.get())
    except:
        messagebox.showerror("Enter numeric value")
    id_da = int(id_lnlE.get())
    id_d = id_da-1219
    if id_d>max:
        name_lbl=Label(booking_frame, text="DONOR ID NOT AVAILABLE ENTER VALID ID",font= ("look", 12), bg = "white")
        name_lbl.place(x=170,y=335)
    else:           
        name = sh1.cell(id_d, 2).value
        print(name)
        BGg = sh1.cell(id_d, 3).value
        age = sh1.cell(id_d, 5).value
        phnum = sh1.cell(id_d, 7).value
        name_lbl=Label(booking_frame, text="Donor Name",font= ("look", 12), bg = "white")
        name_lbl.place(x=170,y=335)
        name_lnlE=Entry(booking_frame,textvariable = name, font=("look","12"), bg="light blue", bd=1)
        name_lnlE.insert(0, name)
        name_lnlE.place(x=350, y=335, width=200, height=40)
        BG_lbl=Label(booking_frame, text="Donor BG",font= ("look", 12), bg = "white")
        BG_lbl.place(x=170,y=400)
        BG_lnlE=Entry(booking_frame,textvariable = BGg, font=("look","12"), bg="light blue", bd=1)
        BG_lnlE.insert(0, BGg)
        BG_lnlE.place(x=350, y=400, width=200, height=40)
        age_lbl=Label(booking_frame, text="Donor Age",font= ("look", 12), bg = "white")
        age_lbl.place(x=170,y=460)
        age_lnlE=Entry(booking_frame,textvariable = age, font=("look","12"), bg="light blue", bd=1)
        age_lnlE.insert(0, age)
        age_lnlE.place(x=350, y=460, width=200, height=40)
        phnum_lbl=Label(booking_frame, text="Donor Phone Number",font= ("look", 12), bg = "white")
        phnum_lbl.place(x=170,y=520)
        phnum_lnlE=Entry(booking_frame,textvariable= phnum, font=("look","12"), bg="light blue", bd=1)
        phnum_lnlE.insert(0, phnum)
        phnum_lnlE.place(x=350, y=520, width=200, height=40)


search_btn = Button(booking_frame, text="Search", width = 20,height= 2,font=("look","14"), bg = "white", fg="black", activebackground="white",
                   command=search_DD)
search_btn.place(x=200, y=200)

root.mainloop()