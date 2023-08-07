from queue import PriorityQueue
from openpyxl import load_workbook

Op = PriorityQueue()
On = PriorityQueue()
Ap = PriorityQueue()
An = PriorityQueue()
Bp = PriorityQueue()
Bn = PriorityQueue()
ABp = PriorityQueue()
ABn = PriorityQueue()

filename = "E:\\ADS PROJECT PYTHON\\Program files\\Donor_details.xlsx"
wb = load_workbook(filename)
sh1 = wb['Donor Details']
a = sh1.max_row
b = sh1.max_column
def roll():
    for i in range(2, a+1):
        x = sh1.cell(i,3).value
        y = globals()[x]
        y.put((sh1.cell(i,5).value, sh1.cell(i,1).value))

def pqdis(var):
    t=list(var.get())
    return t



