from cProfile import label
from tkinter import *
from tokenize import Name
from PIL import Image,ImageTk
from tkinter import messagebox
from matplotlib.pyplot import get
from numpy import append
from tkcalendar import Calendar
from openpyxl import load_workbook

root = Tk()
root.title("Blood Bank Mgmt")
w=1200
h=750
sw = root.winfo_screenwidth()
sh = root.winfo_screenheight()
x = (sw/2) - (w/2)
y = (sh/2) - (h/2)
root.geometry("%dx%d+%d+%d" % (w,h,x,y))

#font style
look = "Adobe Caslon Pro"

#bg

#############bg image


######booking details
booking_frame = Frame(root, bg="white", width=900, height=600).place(x=150, y=50)
title = Label(booking_frame, text= "DONOR REGISTRATION",font= ("look", 20),fg= "red", background = "white")
title.place(x=150, y=90)

######donor_name
name_lbl=Label(booking_frame, text="Enter Donor Name",font= ("look", 12), bg = "white")
name_lbl.place(x=170,y=140)
name_lnlE=Entry(booking_frame,font=("look","12"), bg="light blue", bd=1)
name_lnlE.place(x=350, y=135, width=200, height=40)
######sex_donor
sex_lbl = Label(booking_frame, text="Sex",font=("look", 12), bg="white")
sex_lbl.place(x=170, y=260)
sexa_lbl=["Male","Female"]
clicked1 = StringVar()
clicked1.set("Select Sex")
sexdropbox = OptionMenu(booking_frame, clicked1, *sexa_lbl)
sexdropbox.place(x=350, y=260, width=200, height=40)
sexdropbox.config(bg = "lightblue", activebackground="white")

###Blood group
BG_lbl = Label(booking_frame, text="Blood Group",font=("look", 12), bg="white")
BG_lbl.place(x=170, y=200)
BG_lbl=["ABp", "ABn", "Ap", "An", "Bp","Bn", "Op", "On"]
clicked = StringVar()
clicked.set("Select Blood group")
BGdropbox = OptionMenu(booking_frame, clicked, *BG_lbl)
BGdropbox.place(x=350, y=200, width=200, height=40)
BGdropbox.config(bg = "lightblue", activebackground="white")

###Phone number
Phn_lbl=Label(booking_frame, text="Enter Phone Number",font= ("look", 12), bg = "white")
Phn_lbl.place(x=170,y=320)
Phn_lnlE=Entry(booking_frame,font=("look","12"), bg="light blue", bd=1)
Phn_lnlE.place(x=350, y=320, width=200, height=40)

##Age
age_lbl=Label(booking_frame, text="Enter Age",font= ("look", 12), bg = "white")
age_lbl.place(x=170,y=380)
age_lnlE=Entry(booking_frame,font=("look","12"), bg="light blue", bd=1)
age_lnlE.place(x=350, y=380, width=200, height=40)

##Location
Loc_lbl=Label(booking_frame, text="Enter Location",font= ("look", 12), bg = "white")
Loc_lbl.place(x=170,y=440)
Loc_lnlE=Entry(booking_frame,font=("look","12"), bg="light blue", bd=1)
Loc_lnlE.place(x=350, y=440, width=200, height=40)



######buttons#####
#submit_button
def book_confirm_message():
    D_name = name_lnlE.get()
    BG_name=clicked.get()
    Sex_name = clicked1.get()
    D_loc = Loc_lnlE.get()
    D_age=age_lnlE.get()
    try:
        int(age_lnlE.get())
    except:
        messagebox.showerror("Enter numeric value")
    D_phn = Phn_lnlE.get()
    try:
        int(Phn_lnlE.get())
    except:
        messagebox.showerror("Enter numeric value")
    filename = "E:\\ADS PROJECT PYTHON\\Program files\\Donor_details.xlsx"
    wb = load_workbook(filename)
    sh1 = wb['Donor Details']
    a = sh1.max_row
    b = sh1.max_column
    d_id = sh1.cell(a, 1).value
    cd_id = d_id+1
    print(cd_id)
    sh1.cell(row = a+1, column=1, value = cd_id)
    sh1.cell(row= a+1, column=2, value= D_name)
    sh1.cell(row=a+1, column = 3, value=BG_name)
    sh1.cell(row=a+1, column = 4, value=Sex_name)
    sh1.cell(row=a+1, column = 5, value=D_age)
    sh1.cell(row=a+1, column = 6, value=D_loc)
    sh1.cell(row=a+1, column = 7, value=D_phn)
    wb.save("E:\\ADS PROJECT PYTHON\\Program files\\Donor_details.xlsx")

   
submit_btn = Button(booking_frame, text="Submit",font=("look","12"), bg = "white", activebackground="white",fg="black",
                    command=book_confirm_message)
submit_btn.place(x=450, y=600)
#back_button
def booking_back():
    root.destroy()
    import Menupage
back_btn = Button(booking_frame, text="Back",font=("look","12"), bg = "white", activebackground="white",fg="black",
                  command=booking_back)
back_btn.place(x=650, y=600)

root.mainloop()